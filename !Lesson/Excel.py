from openpyxl import load_workbook
# Указывание существующего файла
fn = "test.xlsx"

wb = load_workbook(fn)
ws = wb["Лист1"]
ws["A5"] = "sfsdf"
wb.save(fn)
wb.close()